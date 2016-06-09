import pandas as pd
import xlrd
import dummy_thread
import openpyxl
import re
import os
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import logging
import matplotlib.pyplot as plt
from matplotlib.dates import date2num
import datetime
import sqlite3

logger = logging.getLogger()
logger.setLevel(logging.WARNING)
logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s')

conn = sqlite3.connect('secfinance.sqlite')
curr = conn.cursor()

curr.executescript('''

DROP TABLE IF EXISTS company;
DROP TABLE IF EXISTS valuetype;
DROP TABLE IF EXISTS summary;
DROP TABLE IF EXISTS dollardenomination;

CREATE TABLE company (
    id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE,
    name TEXT UNIQUE
);

CREATE TABLE valuetype (
    id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE,
    name TEXT UNIQUE
);

CREATE TABLE dollardenomination (
    id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE,
    denomination INTEGER UNIQUE);

CREATE TABLE summary (
    company_id INTEGER,
    valuetype_id INTEGER,
    date TEXT,
    dollaramt INTEGER,
    dollardenomination_id INTEGER)
''')

def extractandstore(namein,valuein,matchlist,tickerin,datein,doldeno):
    for entry in matchlist:
        print 'entry is',entry
        if fuzz.ratio(entry, namein) > 70:
            # TODO Add a check for blank cells
            logger.info('%s,%s', name, fuzz.ratio(entry, name))
            netrevenue = valuein
            curr.execute('''INSERT OR IGNORE INTO company (name)
                VALUES(?)''', (tickerin,))
            curr.execute('''SELECT id FROM company WHERE name = ?''', (tickerin,))
            company_id = curr.fetchone()[0]
            curr.execute(''' INSERT OR IGNORE INTO valuetype (name)
                        VALUES(?)''', ('netrevenue',))
            curr.execute('''SELECT id FROM valuetype WHERE name = ?''', ('netrevenue',))
            valuetype_id = curr.fetchone()[0]
            # denomination table
            curr.execute('''INSERT OR IGNORE INTO dollardenomination(denomination)
                        VALUES(?)''', (doldeno,))
            curr.execute('''SELECT id FROM dollardenomination WHERE denomination = ?''', (doldeno,))
            dollardenomination_id = curr.fetchone()[
                0]  # fetching first column since that is the id that can be used as refernece in main table

            curr.execute('''INSERT OR IGNORE INTO summary (company_id, valuetype_id, date, dollaramt, dollardenomination_id)
            VALUES(?,?,?,?,?)''', (company_id, valuetype_id, datein, netrevenue, dollardenomination_id))
            print 'Netrevenue is ',netrevenue
            break



def xlparse(tickerin,filepath,classname,datein):
    wb = openpyxl.load_workbook(filepath)
    sheetnames = wb.get_sheet_names() #getting list of all sheetnames

    #section for dealing with sheet 1 which is consolidated statement of income
    sheet = wb.get_sheet_by_name(sheetnames[1])
    cola = [] #create new list to store all values in column A
    colb = []
    for cellobj in sheet.columns[0]: #iterate through all values where column0*header column is true..
        cola.append(cellobj.value) #append them to list of columna so we can extract info out of it..

    for cellobj in sheet.columns[1]: #iterate through all values in columnB and store in colb
        colb.append(cellobj.value)

    #extracting information on millions vs thousands from columnArowA
    #format 'Consolidated Condensed Statements of Income - USD ($) shares in Millions, $ in Millions'

    sharedeno = re.findall('shares in ([a-zA-Z]+)',cola[0])
    dollardeno = re.findall('\$ in ([a-zA-Z]+)',cola[0]) #note the use of special character '\' to match for $ vs match at end of line..
    if dollardeno[0] == 'Thousands':
        doldeno = 1000
    elif dollardeno[0] == 'Millions':
        doldeno = 1000000
    else:
        logger.error('Unknown denomination for dollar value')

    print sharedeno[0],dollardeno[0] #regular expression returns a list


    #netrevenue  -cost ofsales = grossmargin 
    #grossmargin - operating expenses = operating income
    #operating expenses = randd + marketing + restructuring + amoritization
    #operating income -   gains(losses) on equity investments + iterests and other... = income before taxes
    #income before taxes - provision for taxes = net incomde
    #netincome/basic shares = basic earnings per share 
    #netincome/diluted shares = diluted earnings per share of commone stock

    netrevenuelist = ['netrevenue','Net sales','Netsales','netsales','Net revenue'] #we will want to keep expanding this list based on how many different variants from different companies
    netincomelist = ['netincome','Net income'] #we will want to keep expanding this list based on how many different variants from different companies

    listoflists = [netrevenuelist,netincomelist]
    logger.debug('%s,%s',cola,colb)
    for name, value in zip(cola,
                           colb):  # looping through both cola and b at the same time .. maybe not efficient time wise and better to use index ?# ?
        #extractandstore(name,value,netrevenuelist,tickerin,datein,doldeno)
        for listentry in listoflists:
            for entry in listentry:
                if fuzz.ratio(entry, name) > 70:
                    # TODO Add a check for blank cells
                    logger.info('%s,%s', name, fuzz.ratio(entry, name))
    #                listentry[0] = value  #first entry of listentry
                    curr.execute('''INSERT OR IGNORE INTO company (name)
                    VALUES(?)''', (tickerin,))
                    curr.execute('''SELECT id FROM company WHERE name = ?''', (tickerin,))
                    company_id = curr.fetchone()[0]
                    curr.execute(''' INSERT OR IGNORE INTO valuetype (name)
                                            VALUES(?)''', (listentry[0],))
                    curr.execute('''SELECT id FROM valuetype WHERE name = ?''', (listentry[0],))
                    valuetype_id = curr.fetchone()[0]
                    # denomination table
                    curr.execute('''INSERT OR IGNORE INTO dollardenomination(denomination)
                            VALUES(?)''', (doldeno,))
                    curr.execute('''SELECT id FROM dollardenomination WHERE denomination = ?''', (doldeno,))
                    dollardenomination_id = curr.fetchone()[
                        0]  # fetching first column since that is the id that can be used as refernece in main table

                    curr.execute('''INSERT OR IGNORE INTO summary (company_id, valuetype_id, date, dollaramt, dollardenomination_id)
                VALUES(?,?,?,?,?)''', (company_id, valuetype_id, datein, value, dollardenomination_id))

                    break
        #for entry in netincomelist:
         #   if fuzz.ratio(entry, name) > 70:
                # TODO Add a check for blank cells
                #logger.info('%s,%s', name, fuzz.ratio(entry, name))
                #netincome = value
                #              curr.execute ('''INSERT OR IGNORE INTO company (name)
                #            VALUES(?)''',(ticker,))
                #                curr.execute (''' INSERT OR IGNORE INTO valuetype (name)
                #                    VALUES(?)''',('netincome',))
                #break
    #print netrevenue, netincome, datein

    conn.commit()

    return
# Main Function
ticker = raw_input("Enter stock ticker")
datelist = []
# parse all files and create a class for each one of those with data
currentpath = os.getcwd()  #getting current directory of .py script
newpath = currentpath + '\\Tickers\\' + ticker.upper() #planning to create new directory with ticker name in upper case
#if os.path.exists(newpath): #check on if path alrerady exists
for root,dirs,files in os.walk(newpath): #walk returns  root path, directories and then the file names
    for name in files:
        #print name
        filepath = (os.path.join(root, name))
        filename =  name.rstrip('.xls') #removing the .xls extension
        date = re.findall(('\d{4}-\d{2}-\d{2}'), filename)
        datet = datetime.datetime.strptime(date[0], '%Y-%m-%d')
        datelist.append(datet)
        logger.debug('%s %s',filepath,filename)
#filepath = 'C:\Users\Aravind\Dropbox\Learning\Programming\Python\Python Fun\SEC_10k_q\AAPL_10-K_2015-10-28.xlsx'
        print filename
        xlparse(ticker, filepath, filename, date[0])


#TODO understand file error with older xls files
#expand to other rows in income sheet
#start plotting
#list = []
#print date2num(datelist)
##for (date,value) in datelist:
##    x = [date2num(datelist)]
##    print x

#for entry in classlist:
#    list.append(entry.netincome)

#for x,y in zip(datelist,list):
#    print x,y

#plt.bar(datelist,list,color="red",linestyle='-',linewidth=5)
#plt.show()