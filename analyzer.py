import pandas as pd
import xlrd
import openpyxl
import re
import os
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import logging

logger = logging.getLogger()
logger.setLevel(logging.WARNING)
logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s')

class quarterlyincome(object):
    
    def __init__(self,innetrevenue,innetincome):
        self.netincome = innetincome
        self.netrevenue = innetrevenue
        
    def test(self):
        print 'testing'
        print  self.netincome
        

def xlparse(filepath,classname):
    wb = openpyxl.load_workbook(filepath)
    sheetnames = wb.get_sheet_names() #getting list of all sheetnames

    #section for dealing with sheet 1 which is consolidated statement of income
    sheet = wb.get_sheet_by_name(sheetnames[1])
    cola = [] #create new list to store all values in column A
    colb = []
    for cellobj in sheet.columns[0]: #iterate through all values where column0*header column is true..
        cola.append(cellobj.value) #append them to list of columna so we can extract info out of it..

    for cellobj in sheet.columns[1]: #iterate through all values in columnB and store in colb
        colb.append(cellobj.value)

    #extracting information on millions vs thousands from columnArowA
    #format 'Consolidated Condensed Statements of Income - USD ($) shares in Millions, $ in Millions'

    sharedeno = re.findall('shares in ([a-zA-Z]+)',cola[0])
    dollardeno = re.findall('\$ in ([a-zA-Z]+)',cola[0]) #note the use of special character '\' to match for $ vs match at end of line..
    #print sharedeno[0],dollardeno[0] #regular expression returns a list


    #netrevenue  -cost ofsales = grossmargin 
    #grossmargin - operating expenses = operating income
    #operating expenses = randd + marketing + restructuring + amoritization
    #operating income -   gains(losses) on equity investments + iterests and other... = income before taxes
    #income before taxes - provision for taxes = net incomde
    #netincome/basic shares = basic earnings per share 
    #netincome/diluted shares = diluted earnings per share of commone stock

    netrevenuelist = ['Net sales','Netsales','netsales','Net revenue','netrevenue'] #we will want to keep expanding this list based on how many different variants from different companies                                                               
    netincomelist = ['Net income','netincome'] #we will want to keep expanding this list based on how many different variants from different companies         

    logger.debug('%s,%s',cola,colb)
    for name,value in zip(cola,colb): #looping through both cola and b at the same time .. maybe not efficient time wise and better to use index ??
        for entry in netrevenuelist:
            if fuzz.ratio(entry,name) > 70:
                #TODO Add a check for blank cells
                logger.info('%s,%s',name,fuzz.ratio(entry,name))
                netrevenue = value
                break
        for entry in netincomelist:
            if fuzz.ratio(entry,name) > 70:
                #TODO Add a check for blank cells
                logger.info('%s,%s',name,fuzz.ratio(entry,name))
                netincome = value
                break
            
    #print netrevenue,netincome
    classname = quarterlyincome(netrevenue,netincome)    #creating a new class with income vlaues 
    print 'Net Revenue  and Net income is',classname.netrevenue,classname.netincome


# parse all files and create a class for each one of those with data
currentpath = os.getcwd()  #getting current directory of .py script
newpath = currentpath + '\\Tickers\\' #planning to create new directory with ticker name in upper case
#if os.path.exists(newpath): #check on if path alrerady exists
for root,dirs,files in os.walk(newpath): #walk returns  root path, directories and then the file names
    for name in files:
        #print name
        filepath = (os.path.join(root, name))
        filename =  name.rstrip('.xls') #removing the .xls extension
        logger.debug('%s %s',filepath,filename)
#filepath = 'C:\Users\Aravind\Dropbox\Learning\Programming\Python\Python Fun\SEC_10k_q\AAPL_10-K_2015-10-28.xlsx'
        print filename
        xlparse(filepath,filename)
        